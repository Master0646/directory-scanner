#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
目录文件扫描器
功能：扫描指定目录，生成目录树，支持导出为Excel或CSV格式
作者：张牛牛
"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime
import threading
from pathlib import Path
import fnmatch
from collections import defaultdict
import subprocess
import json
from openpyxl import Workbook
from openpyxl.styles import Font

class DirectoryScanner:
    """目录扫描器主类"""
    
    def __init__(self, root):
        """初始化GUI界面"""
        self.root = root
        self.root.title("目录文件扫描器 v2.0")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # 设置窗口图标（如果有的话）
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass
        
        # 数据存储
        self.selected_directory = ""
        self.directory_data = []
        self.file_stats = defaultdict(int)
        
        # 扫描设置
        self.max_depth = tk.IntVar(value=3)
        self.include_hidden = tk.BooleanVar(value=False)
        self.show_empty_folders = tk.BooleanVar(value=False)
        self.file_extensions = tk.StringVar(value="*")
        self.display_mode = tk.StringVar(value="detail")
        
        # 排序状态
        self.tree_sort_column = None
        self.tree_sort_reverse = False
        self.detail_sort_column = None
        self.detail_sort_reverse = False
        
        # 扫描状态
        self.scan_stopped = False
        
        # 扫描日志
        self.scan_logs = []
        
        # 创建GUI组件
        self.create_widgets()
        
    def create_widgets(self):
        """创建GUI组件"""
        # 创建笔记本控件（标签页）
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 主扫描页面
        self.main_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.main_frame, text="文件扫描")
        
        # 统计分析页面
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="统计分析")
        
        # 扫描日志页面
        self.logs_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.logs_frame, text="扫描日志")
        
        # 关于页面
        self.about_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.about_frame, text="关于")
        
        # 创建各个页面的内容
        self.create_main_page()
        self.create_stats_page()
        self.create_logs_page()
        self.create_about_page()
        
    def setup_button_style(self):
        """设置按钮样式"""
        # 创建自定义样式
        style = ttk.Style()
        
        # 设置扫描按钮的样式
        style.configure('Scan.TButton', 
                       font=('Arial', 12, 'bold'),
                       padding=(15, 10))
        
        # 设置停止按钮的样式（红色主题）
        style.configure('Stop.TButton',
                       font=('Arial', 12, 'bold'),
                       padding=(15, 10))
        
        # 应用样式到扫描按钮
        self.scan_button.configure(style='Scan.TButton')
        
    def create_main_page(self):
        """创建主扫描页面"""
        # 配置网格权重
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.rowconfigure(4, weight=1)
        
        # 目录选择区域
        ttk.Label(self.main_frame, text="选择目录：", padding="5").grid(row=1, column=0, sticky=tk.W, pady=5)
        
        self.directory_var = tk.StringVar()
        self.directory_entry = ttk.Entry(self.main_frame, textvariable=self.directory_var, width=50)
        self.directory_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=5)
        
        self.browse_button = ttk.Button(self.main_frame, text="浏览", command=self.browse_directory)
        self.browse_button.grid(row=1, column=2, padx=(5, 0), pady=5)
        
        # 扫描设置区域
        scan_settings_frame = ttk.LabelFrame(self.main_frame, padding="5")
        scan_settings_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        scan_settings_frame.columnconfigure(1, weight=1)
        scan_settings_frame.columnconfigure(3, weight=1)
        
        # 第一行：扫描深度和隐藏文件
        ttk.Label(scan_settings_frame, text="扫描深度：").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        
        depth_frame = ttk.Frame(scan_settings_frame)
        depth_frame.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 20))
        
        # 使用Spinbox替代Scale，允许直接输入数字
        self.depth_spinbox = ttk.Spinbox(depth_frame, from_=1, to=50, textvariable=self.max_depth, width=8, validate="key")
        self.depth_spinbox.pack(side=tk.LEFT)
        
        ttk.Label(depth_frame, text="层").pack(side=tk.LEFT, padx=(5, 0))
        
        # 配置验证函数，确保只能输入数字
        vcmd = (self.root.register(self.validate_depth_input), '%P')
        self.depth_spinbox.configure(validatecommand=vcmd)
        
        ttk.Checkbutton(scan_settings_frame, text="包含隐藏文件", variable=self.include_hidden).grid(row=0, column=2, sticky=tk.W, padx=(0, 20))
        
        # 第三行：文件类型快速选择
        ttk.Label(scan_settings_frame, text="文件类型：").grid(row=2, column=0, sticky=tk.W, pady=(10, 0), padx=(0, 5))
        
        type_frame = ttk.Frame(scan_settings_frame)
        type_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        type_frame.columnconfigure(0, weight=1)
        
        # 快速选择按钮
        quick_select_frame = ttk.Frame(type_frame)
        quick_select_frame.grid(row=0, column=0, sticky=tk.W)

        self.quick_filter = tk.StringVar(value="all")
        ttk.Radiobutton(quick_select_frame, text="全部", variable=self.quick_filter, value="all", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="文档", variable=self.quick_filter, value="docs", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="演示", variable=self.quick_filter, value="presentation", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="表格", variable=self.quick_filter, value="spreadsheet", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="图片", variable=self.quick_filter, value="images", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="音频", variable=self.quick_filter, value="audio", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="视频", variable=self.quick_filter, value="video", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="压缩", variable=self.quick_filter, value="archive", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(quick_select_frame, text="代码", variable=self.quick_filter, value="code", command=self.apply_quick_filter).pack(side=tk.LEFT, padx=(0, 10))
        
        # 自定义扩展名输入
        custom_frame = ttk.Frame(type_frame)
        custom_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        custom_frame.columnconfigure(0, weight=1)
        
        self.extension_entry = ttk.Entry(custom_frame, textvariable=self.file_extensions)
        self.extension_entry.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # 绑定事件，当用户输入时自动应用
        self.extension_entry.bind('<KeyRelease>', self.on_extension_change)
        self.extension_entry.bind('<FocusOut>', self.on_extension_change)

        
        # 主要操作按钮区域（右侧显眼位置）
        action_frame = ttk.Frame(scan_settings_frame)
        action_frame.grid(row=0, column=3, rowspan=4, sticky=(tk.N, tk.S), padx=(30, 0))
        
        # 扫描控制按钮（开始/停止切换）
        self.scan_button = ttk.Button(action_frame, text="🚀 开始扫描", command=self.toggle_scan, width=15)
        self.scan_button.pack(side=tk.TOP)
                
        # 为按钮添加样式
        self.setup_button_style()
        
        # 扫描状态标记
        self.is_scanning = False
        
        # 显示模式选择区域
        display_frame = ttk.Frame(self.main_frame)
        display_frame.grid(row=3, column=0, columnspan=3, pady=(10, 5), sticky=tk.W)
        
        ttk.Label(display_frame, text="显示模式：").pack(side=tk.LEFT, padx=(5, 10))

        ttk.Radiobutton(display_frame, text="📊 详细视图", variable=self.display_mode, value="detail", command=self.change_display_mode).pack(side=tk.LEFT, padx=(0, 15))
        ttk.Radiobutton(display_frame, text="📋 列表视图", variable=self.display_mode, value="list", command=self.change_display_mode).pack(side=tk.LEFT, padx=(0, 15))
        ttk.Radiobutton(display_frame, text="🌳 树形视图", variable=self.display_mode, value="tree", command=self.change_display_mode).pack(side=tk.LEFT, padx=(0, 15))
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(self.main_frame, padding="5")
        result_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(1, weight=1)
        
        # 导出按钮区域（右上角）
        export_frame = ttk.Frame(result_frame)
        export_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        export_frame.columnconfigure(0, weight=1)
        
        # 创建导出下拉菜单按钮
        self.export_menubutton = ttk.Menubutton(export_frame, text="📤 导出数据", state="disabled")
        self.export_menubutton.pack(side=tk.RIGHT)
        
        # 创建下拉菜单
        self.export_menu = tk.Menu(self.export_menubutton, tearoff=0)
        self.export_menu.add_command(label="📊 导出为Excel", command=self.export_to_excel)
        self.export_menu.add_command(label="📋 导出为CSV", command=self.export_to_csv)
        self.export_menu.add_command(label="📄 导出为JSON", command=self.export_to_json)
        
        # 配置菜单按钮
        self.export_menubutton.config(menu=self.export_menu)
        
        # 创建显示控件容器
        self.display_container = ttk.Frame(result_frame)
        self.display_container.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.display_container.columnconfigure(0, weight=1)
        self.display_container.rowconfigure(0, weight=1)
        
        # 创建不同的显示控件
        self.create_tree_view()
        self.create_list_view()
        self.create_detail_view()
        self.create_scanning_view()
        
        # 默认显示树形视图
        self.current_view = "detail"
        self.show_view("detail")
        
        # 进度信息区域（右下角）
        progress_frame = ttk.Frame(self.main_frame)
        progress_frame.grid(row=5, column=2, sticky=tk.E, pady=(10, 0))
        
        # 进度信息标签（右对齐）
        self.progress_var = tk.StringVar(value="准备就绪")
        self.progress_label = ttk.Label(progress_frame, textvariable=self.progress_var, font=("Arial", 12), foreground="#666666")
        self.progress_label.pack(side=tk.RIGHT)
        
        # 添加进度条（初始隐藏）
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate', length=200)
        self.progress_bar.pack(side=tk.RIGHT, padx=(0, 10))
        self.progress_bar.pack_forget()  # 初始隐藏
        
    def create_tree_view(self):
        """创建树形视图"""
        self.tree_frame = ttk.Frame(self.display_container)
        
        # 创建Treeview显示目录结构
        self.tree = ttk.Treeview(self.tree_frame, columns=("类型", "大小", "修改时间"), show="tree headings")
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 设置列标题并添加排序功能
        self.tree.heading("#0", text="文件/文件夹名称 ↕", command=lambda: self.sort_tree_column("#0"))
        self.tree.heading("类型", text="类型 ↕", command=lambda: self.sort_tree_column("类型"))
        self.tree.heading("大小", text="大小 ↕", command=lambda: self.sort_tree_column("大小"))
        self.tree.heading("修改时间", text="修改时间 ↕", command=lambda: self.sort_tree_column("修改时间"))
        
        # 设置列宽
        self.tree.column("#0", width=300)
        self.tree.column("类型", width=80)
        self.tree.column("大小", width=100)
        self.tree.column("修改时间", width=150)
        
        # 添加滚动条
        tree_scrollbar_v = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        tree_scrollbar_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=tree_scrollbar_v.set)
        
        tree_scrollbar_h = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        tree_scrollbar_h.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.tree.configure(xscrollcommand=tree_scrollbar_h.set)
        
        self.tree_frame.columnconfigure(0, weight=1)
        self.tree_frame.rowconfigure(0, weight=1)
        
    def create_list_view(self):
        """创建列表视图"""
        self.list_frame = ttk.Frame(self.display_container)
        
        # 创建Listbox显示文件列表
        list_container = ttk.Frame(self.list_frame)
        list_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.file_listbox = tk.Listbox(list_container, font=("Consolas", 10))
        self.file_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 添加滚动条
        list_scrollbar_v = ttk.Scrollbar(list_container, orient="vertical", command=self.file_listbox.yview)
        list_scrollbar_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.file_listbox.configure(yscrollcommand=list_scrollbar_v.set)
        
        list_scrollbar_h = ttk.Scrollbar(list_container, orient="horizontal", command=self.file_listbox.xview)
        list_scrollbar_h.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.file_listbox.configure(xscrollcommand=list_scrollbar_h.set)
        
        list_container.columnconfigure(0, weight=1)
        list_container.rowconfigure(0, weight=1)
        self.list_frame.columnconfigure(0, weight=1)
        self.list_frame.rowconfigure(0, weight=1)
        
    def create_detail_view(self):
        """创建详细视图"""
        self.detail_frame = ttk.Frame(self.display_container)
        
        # 创建详细信息显示区域
        self.detail_tree = ttk.Treeview(self.detail_frame, columns=("路径", "类型", "大小", "修改时间", "扩展名"), show="headings")
        self.detail_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 设置列标题并添加排序功能
        self.detail_tree.heading("路径", text="完整路径 ↕", command=lambda: self.sort_detail_column("路径"))
        self.detail_tree.heading("类型", text="类型 ↕", command=lambda: self.sort_detail_column("类型"))
        self.detail_tree.heading("大小", text="大小 ↕", command=lambda: self.sort_detail_column("大小"))
        self.detail_tree.heading("修改时间", text="修改时间 ↕", command=lambda: self.sort_detail_column("修改时间"))
        self.detail_tree.heading("扩展名", text="扩展名 ↕", command=lambda: self.sort_detail_column("扩展名"))
        
        # 设置列宽
        self.detail_tree.column("路径", width=400)
        self.detail_tree.column("类型", width=80)
        self.detail_tree.column("大小", width=100)
        self.detail_tree.column("修改时间", width=150)
        self.detail_tree.column("扩展名", width=80)
        
        # 添加滚动条
        detail_scrollbar_v = ttk.Scrollbar(self.detail_frame, orient="vertical", command=self.detail_tree.yview)
        detail_scrollbar_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.detail_tree.configure(yscrollcommand=detail_scrollbar_v.set)
        
        detail_scrollbar_h = ttk.Scrollbar(self.detail_frame, orient="horizontal", command=self.detail_tree.xview)
        detail_scrollbar_h.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.detail_tree.configure(xscrollcommand=detail_scrollbar_h.set)
        
        self.detail_frame.columnconfigure(0, weight=1)
        self.detail_frame.rowconfigure(0, weight=1)
        
    def create_scanning_view(self):
        """创建扫描进度显示视图"""
        self.scanning_frame = ttk.Frame(self.display_container)
        
        # 创建居中的进度显示容器
        progress_container = ttk.Frame(self.scanning_frame)
        progress_container.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        # 扫描状态标签
        self.scanning_status_var = tk.StringVar(value="正在扫描...")
        scanning_label = ttk.Label(progress_container, textvariable=self.scanning_status_var, 
                                 font=("Arial", 16, "bold"), foreground="#2E86AB")
        scanning_label.pack(pady=(0, 20))
        
        # 大进度条
        self.scanning_progress_bar = ttk.Progressbar(progress_container, mode='indeterminate', length=300, style="TProgressbar")
        self.scanning_progress_bar.pack(pady=(0, 10))
        
        # 当前扫描文件显示
        self.current_file_var = tk.StringVar(value="")
        current_file_label = ttk.Label(progress_container, textvariable=self.current_file_var, 
                                     font=("Arial", 10), foreground="#666666", wraplength=400)
        current_file_label.pack()
        
    def show_view(self, view_type):
        """显示指定类型的视图"""
        # 隐藏所有视图
        for frame in [self.tree_frame, self.list_frame, self.detail_frame, self.scanning_frame]:
            frame.grid_remove()
        
        # 显示指定视图
        if view_type == "tree":
            self.tree_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        elif view_type == "list":
            self.list_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        elif view_type == "detail":
            self.detail_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        elif view_type == "scanning":
            self.scanning_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.current_view = view_type
        
    def change_display_mode(self):
        """切换显示模式"""
        mode = self.display_mode.get()
        self.show_view(mode)
        
        # 如果有数据，重新显示
        if self.directory_data:
            self.update_display()
            

        
    def create_stats_page(self):
        """创建统计分析页面"""
        # 配置网格权重
        self.stats_frame.columnconfigure(0, weight=1)
        self.stats_frame.rowconfigure(1, weight=1)
        
        # 统计信息显示
        stats_info_frame = ttk.LabelFrame(self.stats_frame, text="扫描统计", padding="10")
        stats_info_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(10, 5))
        stats_info_frame.columnconfigure(1, weight=1)
        
        # 基本统计
        self.total_files_var = tk.StringVar(value="0")
        self.total_folders_var = tk.StringVar(value="0")
        self.total_size_var = tk.StringVar(value="0 B")
        
        ttk.Label(stats_info_frame, text="文件总数：").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Label(stats_info_frame, textvariable=self.total_files_var).grid(row=0, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(stats_info_frame, text="文件夹总数：").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Label(stats_info_frame, textvariable=self.total_folders_var).grid(row=1, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(stats_info_frame, text="总大小：").grid(row=2, column=0, sticky=tk.W, pady=2)
        ttk.Label(stats_info_frame, textvariable=self.total_size_var).grid(row=2, column=1, sticky=tk.W, pady=2)
        
        # 文件类型统计
        type_stats_frame = ttk.LabelFrame(self.stats_frame, text="文件类型统计", padding="10")
        type_stats_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        type_stats_frame.columnconfigure(0, weight=1)
        type_stats_frame.rowconfigure(0, weight=1)
        
        # 创建文件类型统计表格
        self.type_tree = ttk.Treeview(type_stats_frame, columns=("数量", "总大小", "平均大小"), show="tree headings")
        self.type_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.type_tree.heading("#0", text="文件类型")
        self.type_tree.heading("数量", text="数量")
        self.type_tree.heading("总大小", text="总大小")
        self.type_tree.heading("平均大小", text="平均大小")
        
        self.type_tree.column("#0", width=150)
        self.type_tree.column("数量", width=80)
        self.type_tree.column("总大小", width=100)
        self.type_tree.column("平均大小", width=100)
        
        # 添加滚动条
        type_scrollbar = ttk.Scrollbar(type_stats_frame, orient="vertical", command=self.type_tree.yview)
        type_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.type_tree.configure(yscrollcommand=type_scrollbar.set)
        
    def browse_directory(self):
        """浏览并选择目录"""
        directory = filedialog.askdirectory(title="选择要扫描的目录")
        if directory:
            self.directory_var.set(directory)
            self.selected_directory = directory
            
    def validate_depth_input(self, value):
        """验证深度输入是否为有效数字"""
        if value == "":
            return True
        try:
            depth = int(value)
            return 1 <= depth <= 50
        except ValueError:
            return False
        
    def apply_quick_filter(self):
        """应用快速文件类型过滤"""
        filter_type = self.quick_filter.get()
        
        if filter_type == "all":
            self.file_extensions.set("*")
        elif filter_type == "docs":
            extensions = ".txt,.doc,.docx,.pdf,.rtf,.odt"
            self.file_extensions.set(extensions)
        elif filter_type == "images":
            extensions = ".jpg,.jpeg,.png,.gif,.bmp,.svg,.tiff"
            self.file_extensions.set(extensions)
        elif filter_type == "audio":
            extensions = ".mp3,.wav,.flac,.aac,.ogg,.m4a"
            self.file_extensions.set(extensions)
        elif filter_type == "video":
            extensions = ".mp4,.avi,.mkv,.mov,.wmv,.flv,.webm"
            self.file_extensions.set(extensions)
        elif filter_type == "archive":
            extensions = ".zip,.rar,.7z,.tar,.gz,.bz2"
            self.file_extensions.set(extensions)
        elif filter_type == "code":
            extensions = ".py,.js,.html,.css,.java,.cpp,.c,.php"
            self.file_extensions.set(extensions)
        elif filter_type == "spreadsheet":
            extensions = ".xls,.xlsx,.csv,.ods"
            self.file_extensions.set(extensions)
        elif filter_type == "presentation":
            extensions = ".ppt,.pptx,.odp"
            self.file_extensions.set(extensions)
    
    def on_extension_change(self, event=None):
        """当扩展名输入框内容改变时自动应用"""
        # 添加短暂延迟，避免频繁触发
        if hasattr(self, '_extension_timer'):
            self.root.after_cancel(self._extension_timer)
        self._extension_timer = self.root.after(500, self.apply_custom_filter)
    
    def apply_custom_filter(self):
        """应用自定义扩展名过滤"""
        extensions = self.file_extensions.get().strip()
        if not extensions:
            extensions = "*"
            self.file_extensions.set(extensions)
        
        if extensions == "*":
            self.quick_filter.set("all")
        else:
            # 清除快速选择
            self.quick_filter.set("")
        

        

            

            

        
    def toggle_scan(self):
        """切换扫描状态（开始/停止）"""
        if self.is_scanning:
            self.stop_scan()
        else:
            self.start_scan()
            
    def update_scan_button_state(self, is_scanning):
        """更新扫描按钮的状态和外观"""
        if is_scanning:
            self.scan_button.configure(
                text="⏹️ 停止扫描",
                style='Stop.TButton'
            )
            self.progress_var.set("正在扫描...")
            # 显示并启动进度条动画
            self.progress_bar.pack(side=tk.RIGHT, padx=(0, 10))
            self.progress_bar.start(10)  # 每10ms更新一次
        else:
            self.scan_button.configure(
                text="🚀 开始扫描",
                style='Scan.TButton'
            )
            # 停止并隐藏进度条
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            
            # 扫描状态指示器已移除
            
            if hasattr(self, 'directory_data') and self.directory_data:
                file_count = len([item for item in self.directory_data if item['类型'] == '文件'])
                folder_count = len([item for item in self.directory_data if item['类型'] == '文件夹'])
                self.progress_var.set(f"✅ 扫描完成 - {file_count} 个文件，{folder_count} 个文件夹")
            else:
                self.progress_var.set("准备就绪")
    
    def stop_scan(self):
        """停止扫描"""
        self.scan_stopped = True
        self.is_scanning = False
        
        # 记录停止扫描
        self.add_log("用户手动停止扫描", "warning")
        
        # 停止扫描进度条并恢复视图
        if hasattr(self, 'scanning_progress_bar'):
            self.scanning_progress_bar.stop()
        if hasattr(self, 'directory_data') and self.directory_data:
            self.show_view("detail")  # 如果有数据则显示详细视图
            # 重新显示右下角的进度信息
            self.progress_label.pack(side=tk.RIGHT)
        
        # 更新按钮状态和外观
        self.update_scan_button_state(False)
        self.progress_var.set("扫描已停止")
        
        # 启用导出按钮
        if hasattr(self, 'directory_data') and self.directory_data:
            self.export_menubutton.config(state="normal")
        
    def should_include_file(self, file_path, file_name):
        """判断是否应该包含此文件"""
        # 检查隐藏文件
        if not self.include_hidden.get() and file_name.startswith('.'):
            return False
            
        # 检查文件扩展名过滤
        extensions = self.file_extensions.get().strip()
        if extensions and extensions != "*":
            file_ext = Path(file_path).suffix.lower()  # 获取文件扩展名（带点，如.txt）
            file_ext_no_dot = file_ext[1:] if file_ext.startswith('.') else file_ext  # 不带点的扩展名（如txt）
            
            # 处理允许的扩展名列表，支持带点和不带点的格式
            allowed_exts = []
            for ext in extensions.split(","):
                ext = ext.strip().lower()
                if ext:
                    # 添加带点和不带点的格式
                    if ext.startswith('.'):
                        allowed_exts.append(ext)  # 带点格式
                        allowed_exts.append(ext[1:])  # 不带点格式
                    else:
                        allowed_exts.append('.' + ext)  # 带点格式
                        allowed_exts.append(ext)  # 不带点格式
            
            # 检查文件扩展名是否在允许列表中
            if allowed_exts and file_ext not in allowed_exts and file_ext_no_dot not in allowed_exts:
                return False
                
        return True
        
    def start_scan(self):
        """开始扫描目录"""
        if not self.directory_var.get():
            messagebox.showwarning("警告", "请先选择要扫描的目录！")
            self.add_log("扫描失败：未选择目录", "warning")
            return
            
        if not os.path.exists(self.directory_var.get()):
            messagebox.showerror("错误", "选择的目录不存在！")
            self.add_log(f"扫描失败：目录不存在 - {self.directory_var.get()}", "error")
            return
            
        # 记录扫描开始
        directory = self.directory_var.get()
        self.add_log(f"开始扫描目录：{directory}", "info")
        self.add_log(f"扫描设置 - 深度：{self.max_depth.get()}层，包含隐藏文件：{'是' if self.include_hidden.get() else '否'}", "info")
        
        # 设置扫描状态
        self.scan_stopped = False
        self.is_scanning = True
        
        # 清空之前的结果
        self.tree.delete(*self.tree.get_children())
        self.file_listbox.delete(0, tk.END)
        self.detail_tree.delete(*self.detail_tree.get_children())
        self.directory_data.clear()
        self.file_stats.clear()
        
        # 更新按钮状态和外观
        self.update_scan_button_state(True)
        
        # 禁用导出按钮
        self.export_menubutton.config(state="disabled")
        
        # 显示扫描进度界面
        self.show_view("scanning")
        self.scanning_progress_bar.start(10)
        self.scanning_status_var.set("正在扫描...")
        self.current_file_var.set("准备开始扫描")
        
        # 隐藏右下角的进度信息
        self.progress_label.pack_forget()
        self.progress_bar.pack_forget()
        
        # 在新线程中执行扫描
        scan_thread = threading.Thread(target=self.scan_directory)
        scan_thread.daemon = True
        scan_thread.start()
        
    def scan_directory(self):
        """扫描目录（在后台线程中执行）"""
        start_time = datetime.now()
        try:
            root_path = self.directory_var.get()
            
            # 插入根目录
            root_item = self.tree.insert("", "end", text=os.path.basename(root_path) or root_path, 
                                        values=("文件夹", "-", self.get_modification_time(root_path)))
            
            # 递归扫描目录
            self.scan_recursive(root_path, root_item, "", 0)
            
            if not self.scan_stopped:
                # 计算扫描统计
                file_count = len([item for item in self.directory_data if item['类型'] == '文件'])
                folder_count = len([item for item in self.directory_data if item['类型'] == '文件夹'])
                end_time = datetime.now()
                duration = (end_time - start_time).total_seconds()
                
                # 记录扫描完成
                self.root.after(0, lambda: self.add_log(f"扫描完成！共发现 {file_count} 个文件，{folder_count} 个文件夹，耗时 {duration:.2f} 秒", "success"))
                
                # 停止扫描进度条并恢复视图
                self.root.after(0, lambda: self.scanning_progress_bar.stop())
                self.root.after(0, lambda: self.show_view("detail"))  # 恢复到详细视图
                
                # 重新显示右下角的进度信息
                self.root.after(0, lambda: self.progress_label.pack(side=tk.RIGHT))
                
                # 更新显示
                self.root.after(0, self.update_display)
                
                # 更新显示
                self.root.after(0, self.update_stats_display)
                
                # 扫描完成，启用导出按钮
                self.root.after(0, lambda: self.export_menubutton.config(state="normal"))
                self.root.after(0, lambda: self.progress_var.set("扫描完成"))
            else:
                # 扫描被停止
                self.root.after(0, lambda: self.add_log("扫描被用户停止", "warning"))
            
        except Exception as e:
            error_msg = f"扫描过程中发生错误：{str(e)}"
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
            self.root.after(0, lambda: self.progress_var.set("扫描失败"))
            self.root.after(0, lambda: self.add_log(error_msg, "error"))
        finally:
            # 恢复按钮状态
            self.is_scanning = False
            self.root.after(0, lambda: self.update_scan_button_state(False))
            
    def scan_recursive(self, path, parent_item, relative_path, current_depth):
        """递归扫描目录"""
        if self.scan_stopped:
            return False
            
        # 检查深度限制
        if current_depth >= self.max_depth.get():
            return False
            
        has_valid_content = False  # 标记是否包含有效内容
        
        try:
            items = os.listdir(path)
            
            # 按名称排序（默认排序方式）
            items.sort()
            
            # 分别处理文件和文件夹
            files = []
            folders = []
            
            for item in items:
                if self.scan_stopped:
                    break
                    
                item_path = os.path.join(path, item)
                
                try:
                    if os.path.isdir(item_path):
                        # 检查隐藏文件夹
                        if not self.include_hidden.get() and item.startswith('.'):
                            continue
                        folders.append((item, item_path))
                    else:
                        # 文件 - 检查是否应该包含
                        if self.should_include_file(item_path, item):
                            files.append((item, item_path))
                            
                except PermissionError:
                    # 记录权限错误
                    self.root.after(0, lambda p=item_path: self.add_log(f"权限不足，无法访问：{p}", "warning"))
                    continue
                except Exception as e:
                    # 记录其他错误
                    self.root.after(0, lambda p=item_path, err=str(e): self.add_log(f"访问文件时出错：{p} - {err}", "error"))
                    continue
            
            # 处理文件
            for item, item_path in files:
                if self.scan_stopped:
                    break
                    
                item_relative_path = os.path.join(relative_path, item) if relative_path else item
                
                file_size = self.get_file_size(item_path)
                mod_time = self.get_modification_time(item_path)
                
                self.tree.insert(parent_item, "end", text=item,
                               values=("文件", file_size, mod_time))
                
                # 添加到数据列表
                self.directory_data.append({
                    "路径": item_relative_path,
                    "名称": item,
                    "类型": "文件",
                    "大小": file_size,
                    "修改时间": mod_time
                })
                
                # 统计文件信息
                self.file_stats["文件"] = self.file_stats.get("文件", 0) + 1
                file_ext = Path(item_path).suffix.lower()
                if file_ext:
                    self.file_stats[file_ext] = self.file_stats.get(file_ext, 0) + 1
                else:
                    self.file_stats["无扩展名"] = self.file_stats.get("无扩展名", 0) + 1
                
                has_valid_content = True
                
                # 更新进度显示（每处理10个文件更新一次，避免过于频繁）
                if len(self.directory_data) % 10 == 0:
                    file_count = len([item for item in self.directory_data if item['类型'] == '文件'])
                    folder_count = len([item for item in self.directory_data if item['类型'] == '文件夹'])
                    # 更新扫描界面的进度信息
                    self.root.after(0, lambda fc=file_count, foc=folder_count: 
                                    self.scanning_status_var.set(f"已扫描 {fc} 个文件，{foc} 个文件夹"))
                
                # 更新当前扫描的文件
                self.root.after(0, lambda f=item_relative_path: 
                                self.current_file_var.set(f"正在扫描：{f}"))
            
            # 处理文件夹
            for item, item_path in folders:
                if self.scan_stopped:
                    break
                    
                item_relative_path = os.path.join(relative_path, item) if relative_path else item
                
                # 创建临时的文件夹项目
                temp_folder_item = self.tree.insert(parent_item, "end", text=item,
                                                   values=("文件夹", "-", self.get_modification_time(item_path)))
                
                # 递归扫描子目录
                folder_has_content = self.scan_recursive(item_path, temp_folder_item, item_relative_path, current_depth + 1)
                
                # 检查文件夹是否应该保留
                if folder_has_content or self.show_empty_folders.get():
                    # 保留文件夹
                    self.directory_data.append({
                        "路径": item_relative_path,
                        "名称": item,
                        "类型": "文件夹",
                        "大小": "-",
                        "修改时间": self.get_modification_time(item_path)
                    })
                    
                    # 统计文件夹数量
                    self.file_stats["文件夹"] = self.file_stats.get("文件夹", 0) + 1
                    has_valid_content = True
                else:
                    # 删除空文件夹项目
                    self.tree.delete(temp_folder_item)
                    
        except PermissionError:
            # 记录权限错误
            self.root.after(0, lambda p=path: self.add_log(f"权限不足，无法访问目录：{p}", "warning"))
        except Exception as e:
            # 记录其他错误
            self.root.after(0, lambda p=path, err=str(e): self.add_log(f"扫描目录时出错：{p} - {err}", "error"))
            
        return has_valid_content
            
    def get_file_size(self, file_path):
        """获取文件大小（格式化显示）"""
        try:
            size = os.path.getsize(file_path)
            if size < 1024:
                return f"{size} B"
            elif size < 1024 * 1024:
                return f"{size / 1024:.1f} KB"
            elif size < 1024 * 1024 * 1024:
                return f"{size / (1024 * 1024):.1f} MB"
            else:
                return f"{size / (1024 * 1024 * 1024):.1f} GB"
        except:
            return "未知"
            
    def get_modification_time(self, file_path):
        """获取文件修改时间"""
        try:
            timestamp = os.path.getmtime(file_path)
            return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
        except:
            return "未知"
            
    def export_to_excel(self):
        """导出为Excel文件（支持文件名超链接）"""
        if not self.directory_data:
            messagebox.showwarning("警告", "没有数据可导出！")
            self.add_log("导出失败：没有数据可导出", "warning")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                import os
                
                # 过滤数据：只导出文件，不导出文件夹，与页面显示保持一致
                files_only_data = []
                for item in self.directory_data:
                    if item['类型'] == '文件':
                        # 添加扩展名列，与详细视图保持一致
                        file_ext = Path(item['名称']).suffix.lower() if item['名称'] else '-'
                        export_item = item.copy()
                        export_item['扩展名'] = file_ext
                        files_only_data.append(export_item)
                
                if not files_only_data:
                    messagebox.showwarning("警告", "没有文件数据可导出！")
                    return
                
                # 创建工作簿和工作表
                wb = Workbook()
                ws = wb.active
                ws.title = "文件列表"
                
                # 设置列标题
                headers = ['路径', '名称', '类型', '大小', '修改时间', '扩展名']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    ws.cell(row=1, column=col).font = Font(bold=True)
                
                # 获取扫描的根目录
                root_directory = self.directory_var.get()
                
                # 填充数据并添加超链接
                for row, item in enumerate(files_only_data, 2):
                    # 构建完整的文件路径
                    full_file_path = os.path.join(root_directory, item['路径'])
                    
                    # 路径列
                    ws.cell(row=row, column=1, value=item['路径'])
                    
                    # 名称列（添加超链接）
                    name_cell = ws.cell(row=row, column=2, value=item['名称'])
                    # 创建文件系统超链接
                    if os.path.exists(full_file_path):
                        # 在Mac系统中使用file://协议，需要三个斜杠
                        file_url = f"file:///{full_file_path}"
                        name_cell.hyperlink = file_url
                        name_cell.font = Font(color="0000FF", underline="single")  # 蓝色下划线
                    
                    # 其他列
                    ws.cell(row=row, column=3, value=item['类型'])
                    ws.cell(row=row, column=4, value=item['大小'])
                    ws.cell(row=row, column=5, value=item['修改时间'])
                    ws.cell(row=row, column=6, value=item['扩展名'])
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 保存文件
                wb.save(file_path)
                self.show_export_success_dialog(
                    file_path, 
                    "Excel文件（含超链接）", 
                    len(files_only_data), 
                    "点击Excel中的文件名可直接打开文件所在目录"
                )
                self.add_log(f"成功导出Excel文件（含超链接）：{file_path}，共 {len(files_only_data)} 条文件记录", "success")
            except Exception as e:
                error_msg = f"导出Excel文件失败：{str(e)}"
                messagebox.showerror("错误", error_msg)
                self.add_log(error_msg, "error")
                
    def export_to_csv(self):
        """导出为CSV文件（包含完整文件路径）"""
        if not self.directory_data:
            messagebox.showwarning("警告", "没有数据可导出！")
            self.add_log("导出失败：没有数据可导出", "warning")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="保存CSV文件",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                # 获取根目录路径
                root_directory = getattr(self, 'current_directory', '')
                
                # 过滤数据：只导出文件，不导出文件夹，与页面显示保持一致
                files_only_data = []
                for item in self.directory_data:
                    if item['类型'] == '文件':
                        # 添加扩展名列，与详细视图保持一致
                        file_ext = Path(item['名称']).suffix.lower() if item['名称'] else '-'
                        # 构建完整文件路径
                        full_file_path = os.path.join(root_directory, item['路径'])
                        
                        export_item = item.copy()
                        export_item['扩展名'] = file_ext
                        export_item['完整路径'] = full_file_path
                        files_only_data.append(export_item)
                
                df = pd.DataFrame(files_only_data)
                # 重新排列列的顺序，添加完整路径列便于用户访问文件
                if not df.empty:
                    df = df[['路径', '名称', '类型', '大小', '修改时间', '扩展名', '完整路径']]
                
                df.to_csv(file_path, index=False, encoding='utf-8-sig')  # 使用utf-8-sig编码支持中文
                self.show_export_success_dialog(
                    file_path, 
                    "CSV文件（含完整路径）", 
                    len(files_only_data), 
                    "CSV中包含完整路径列，可复制路径直接访问文件"
                )
                self.add_log(f"成功导出CSV文件（含完整路径）：{file_path}，共 {len(files_only_data)} 条文件记录", "success")
            except Exception as e:
                error_msg = f"导出CSV文件失败：{str(e)}"
                messagebox.showerror("错误", error_msg)
                self.add_log(error_msg, "error")
                
    def export_to_json(self):
        """导出为JSON文件（包含完整文件路径）"""
        if not self.directory_data:
            messagebox.showwarning("警告", "没有数据可导出！")
            self.add_log("导出失败：没有数据可导出", "warning")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="保存JSON文件",
            defaultextension=".json",
            filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                # 获取根目录路径
                root_directory = getattr(self, 'current_directory', '')
                
                # 过滤数据：只导出文件，不导出文件夹，与页面显示保持一致
                files_only_data = []
                for item in self.directory_data:
                    if item['类型'] == '文件':
                        # 添加扩展名列，与详细视图保持一致
                        file_ext = Path(item['名称']).suffix.lower() if item['名称'] else '-'
                        # 构建完整文件路径
                        full_file_path = os.path.join(root_directory, item['路径'])
                        
                        export_item = item.copy()
                        export_item['扩展名'] = file_ext
                        export_item['完整路径'] = full_file_path
                        files_only_data.append(export_item)
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                
                self.show_export_success_dialog(
                    file_path, 
                    "JSON文件（含完整路径）", 
                    len(files_only_data), 
                    "JSON中包含完整路径字段，可复制路径直接访问文件"
                )
                self.add_log(f"成功导出JSON文件（含完整路径）：{file_path}，共 {len(files_only_data)} 条文件记录", "success")
            except Exception as e:
                error_msg = f"导出JSON文件失败：{str(e)}"
                messagebox.showerror("错误", error_msg)
                self.add_log(error_msg, "error")
                
    def update_display(self):
        """更新显示内容"""
        # 过滤数据：列表视图和详细视图只显示文件，不显示文件夹
        files_only_data = [item for item in self.directory_data if item['类型'] == '文件']
        
        # 更新列表视图（只显示文件）
        self.file_listbox.delete(0, tk.END)
        for item in files_only_data:
            display_text = f"{item['类型']:6} | {item['名称']:30} | {item['大小']:10} | {item['修改时间']}"
            self.file_listbox.insert(tk.END, display_text)
            
        # 更新详细视图（只显示文件）
        for item in self.detail_tree.get_children():
            self.detail_tree.delete(item)
            
        for item in files_only_data:
            file_ext = Path(item['名称']).suffix.lower() if item['类型'] == '文件' else '-'
            self.detail_tree.insert("", "end", 
                                   values=(item['路径'], item['类型'], item['大小'], 
                                          item['修改时间'], file_ext))
                                          
    def update_stats_display(self):
        """更新统计显示"""
        # 更新基本统计
        total_files = self.file_stats.get("文件", 0)
        total_folders = self.file_stats.get("文件夹", 0)
        
        self.total_files_var.set(str(total_files))
        self.total_folders_var.set(str(total_folders))
        
        # 计算总大小
        total_size = 0
        for item in self.directory_data:
            if item['类型'] == '文件' and item['大小'] != '-':
                size_str = item['大小']
                try:
                    if 'KB' in size_str:
                        total_size += float(size_str.replace(' KB', '')) * 1024
                    elif 'MB' in size_str:
                        total_size += float(size_str.replace(' MB', '')) * 1024 * 1024
                    elif 'GB' in size_str:
                        total_size += float(size_str.replace(' GB', '')) * 1024 * 1024 * 1024
                    elif 'B' in size_str:
                        total_size += float(size_str.replace(' B', ''))
                except:
                    pass
                    
        self.total_size_var.set(self.get_file_size(int(total_size)))
        
        # 更新文件类型统计
        for item in self.type_tree.get_children():
            self.type_tree.delete(item)
            
        # 统计文件类型
        type_stats = {}
        for item in self.directory_data:
            if item['类型'] == '文件':
                file_ext = Path(item['名称']).suffix.lower()
                if not file_ext:
                    file_ext = '无扩展名'
                    
                if file_ext not in type_stats:
                    type_stats[file_ext] = {'count': 0, 'total_size': 0}
                    
                type_stats[file_ext]['count'] += 1
                
                # 计算大小
                size_str = item['大小']
                try:
                    if 'KB' in size_str:
                        size_bytes = float(size_str.replace(' KB', '')) * 1024
                    elif 'MB' in size_str:
                        size_bytes = float(size_str.replace(' MB', '')) * 1024 * 1024
                    elif 'GB' in size_str:
                        size_bytes = float(size_str.replace(' GB', '')) * 1024 * 1024 * 1024
                    elif 'B' in size_str:
                        size_bytes = float(size_str.replace(' B', ''))
                    else:
                        size_bytes = 0
                    type_stats[file_ext]['total_size'] += size_bytes
                except:
                    pass
                    
        # 显示文件类型统计
        for file_type, stats in sorted(type_stats.items(), key=lambda x: x[1]['count'], reverse=True):
            count = stats['count']
            total_size = self.get_file_size(int(stats['total_size']))
            avg_size = self.get_file_size(int(stats['total_size'] / count)) if count > 0 else '0 B'
            
            self.type_tree.insert("", "end", text=file_type,
                                 values=(count, total_size, avg_size))
                                 
    def sort_tree_column(self, column):
        """对树形视图的列进行排序"""
        # 切换排序方向
        if self.tree_sort_column == column:
            self.tree_sort_reverse = not self.tree_sort_reverse
        else:
            self.tree_sort_column = column
            self.tree_sort_reverse = False
            
        # 获取所有项目
        items = []
        for child in self.tree.get_children():
            item_data = self.get_tree_item_data(child, column)
            items.append((item_data, child))
            
        # 排序
        items.sort(key=lambda x: x[0], reverse=self.tree_sort_reverse)
        
        # 重新排列项目
        for index, (_, child) in enumerate(items):
            self.tree.move(child, '', index)
            
        # 更新列标题显示排序状态
        self.update_tree_column_headers()
        
    def get_tree_item_data(self, item, column):
        """获取树形视图项目的排序数据"""
        if column == "#0":
            return self.tree.item(item, 'text').lower()
        elif column == "类型":
            return self.tree.set(item, "类型").lower()
        elif column == "大小":
            size_str = self.tree.set(item, "大小")
            return self.parse_size_for_sort(size_str)
        elif column == "修改时间":
            time_str = self.tree.set(item, "修改时间")
            try:
                return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
            except:
                return datetime.min
        return ""
        
    def sort_detail_column(self, column):
        """对详细视图的列进行排序"""
        # 切换排序方向
        if self.detail_sort_column == column:
            self.detail_sort_reverse = not self.detail_sort_reverse
        else:
            self.detail_sort_column = column
            self.detail_sort_reverse = False
            
        # 获取所有项目
        items = []
        for child in self.detail_tree.get_children():
            item_data = self.get_detail_item_data(child, column)
            values = self.detail_tree.item(child, 'values')
            items.append((item_data, values))
            
        # 排序
        items.sort(key=lambda x: x[0], reverse=self.detail_sort_reverse)
        
        # 清空并重新插入
        for child in self.detail_tree.get_children():
            self.detail_tree.delete(child)
            
        for _, values in items:
            self.detail_tree.insert("", "end", values=values)
            
        # 更新列标题显示排序状态
        self.update_detail_column_headers()
        
    def get_detail_item_data(self, item, column):
        """获取详细视图项目的排序数据"""
        values = self.detail_tree.item(item, 'values')
        if column == "路径":
            return values[0].lower()
        elif column == "类型":
            return values[1].lower()
        elif column == "大小":
            return self.parse_size_for_sort(values[2])
        elif column == "修改时间":
            try:
                return datetime.strptime(values[3], "%Y-%m-%d %H:%M:%S")
            except:
                return datetime.min
        elif column == "扩展名":
            return values[4].lower()
        return ""
        
    def parse_size_for_sort(self, size_str):
        """解析大小字符串用于排序"""
        if size_str == "-" or size_str == "未知":
            return 0
        try:
            if 'KB' in size_str:
                return float(size_str.replace(' KB', '')) * 1024
            elif 'MB' in size_str:
                return float(size_str.replace(' MB', '')) * 1024 * 1024
            elif 'GB' in size_str:
                return float(size_str.replace(' GB', '')) * 1024 * 1024 * 1024
            elif 'B' in size_str:
                return float(size_str.replace(' B', ''))
        except:
            pass
        return 0
        
    def update_tree_column_headers(self):
        """更新树形视图列标题显示排序状态"""
        columns = [("#0", "文件/文件夹名称"), ("类型", "类型"), ("大小", "大小"), ("修改时间", "修改时间")]
        
        for col_id, col_name in columns:
            if col_id == self.tree_sort_column:
                arrow = " ↓" if self.tree_sort_reverse else " ↑"
                self.tree.heading(col_id, text=col_name + arrow)
            else:
                self.tree.heading(col_id, text=col_name + " ↕")
                
    def update_detail_column_headers(self):
        """更新详细视图列标题显示排序状态"""
        columns = [("路径", "完整路径"), ("类型", "类型"), ("大小", "大小"), ("修改时间", "修改时间"), ("扩展名", "扩展名")]
        
        for col_id, col_name in columns:
            if col_id == self.detail_sort_column:
                arrow = " ↓" if self.detail_sort_reverse else " ↑"
                self.detail_tree.heading(col_id, text=col_name + arrow)
            else:
                self.detail_tree.heading(col_id, text=col_name + " ↕")
                
    def create_logs_page(self):
        """创建扫描日志页面"""
        # 配置网格权重
        self.logs_frame.columnconfigure(0, weight=1)
        self.logs_frame.rowconfigure(1, weight=1)
        
        # 日志控制区域
        log_control_frame = ttk.Frame(self.logs_frame)
        log_control_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(10, 5))
        log_control_frame.columnconfigure(1, weight=1)
        
        # 清空日志按钮
        self.clear_logs_button = ttk.Button(log_control_frame, text="清空日志", command=self.clear_logs)
        self.clear_logs_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 导出日志按钮
        self.export_logs_button = ttk.Button(log_control_frame, text="导出日志", command=self.export_logs)
        self.export_logs_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 日志统计信息
        self.log_info_var = tk.StringVar(value="日志条数：0")
        self.log_info_label = ttk.Label(log_control_frame, textvariable=self.log_info_var)
        self.log_info_label.pack(side=tk.RIGHT)
        
        # 日志显示区域
        log_display_frame = ttk.LabelFrame(self.logs_frame, text="扫描日志", padding="5")
        log_display_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        log_display_frame.columnconfigure(0, weight=1)
        log_display_frame.rowconfigure(0, weight=1)
        
        # 创建日志文本框
        self.log_text = tk.Text(log_display_frame, wrap=tk.WORD, font=("Consolas", 10))
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 添加滚动条
        log_scrollbar = ttk.Scrollbar(log_display_frame, orient="vertical", command=self.log_text.yview)
        log_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        # 设置日志文本框为只读
        self.log_text.configure(state="disabled")
        
        # 配置日志文本样式
        self.log_text.tag_configure("info", foreground="#2E8B57")
        self.log_text.tag_configure("warning", foreground="#FF8C00")
        self.log_text.tag_configure("error", foreground="#DC143C")
        self.log_text.tag_configure("success", foreground="#228B22")
        
    def add_log(self, message, log_type="info"):
        """添加日志条目
        
        Args:
            message (str): 日志消息
            log_type (str): 日志类型 ('info', 'warning', 'error', 'success')
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        # 添加到日志列表
        self.scan_logs.append({
            "timestamp": timestamp,
            "message": message,
            "type": log_type
        })
        
        # 更新日志显示
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, log_entry, log_type)
        self.log_text.configure(state="disabled")
        self.log_text.see(tk.END)  # 自动滚动到最新日志
        
        # 更新日志统计
        self.log_info_var.set(f"日志条数：{len(self.scan_logs)}")
    
    def show_export_success_dialog(self, file_path, file_type, record_count, tips=""):
        """显示导出成功对话框，支持直接打开文件
        
        Args:
            file_path (str): 导出的文件路径
            file_type (str): 文件类型描述
            record_count (int): 导出的记录数量
            tips (str): 额外提示信息
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("导出成功")
        dialog.geometry("450x260")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 主框架
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # 成功图标和标题
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill="x", pady=(0, 15))
        
        ttk.Label(title_frame, text="✅", font=("Arial", 20)).pack(side="left")
        ttk.Label(title_frame, text="导出成功！", font=("Arial", 14, "bold")).pack(side="left", padx=(10, 0))
        
        # 文件信息
        info_text = f"文件类型：{file_type}\n文件路径：{file_path}\n记录数量：{record_count} 条"
        if tips:
            info_text += f"\n\n💡 {tips}"
        
        ttk.Label(main_frame, text=info_text, wraplength=400, justify="left").pack(fill="x", pady=(0, 20))
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")
        
        # 打开文件按钮
        def open_file():
            try:
                subprocess.call(['open', file_path])
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件：{str(e)}")
        
        ttk.Button(button_frame, text="📂 打开文件", command=open_file).pack(side="left", padx=(0, 10))
        
        # 打开文件夹按钮
        def open_folder():
            try:
                folder_path = os.path.dirname(file_path)
                subprocess.call(['open', folder_path])
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件夹：{str(e)}")
        
        ttk.Button(button_frame, text="📁 打开文件夹", command=open_folder).pack(side="left", padx=(0, 10))
        
        # 关闭按钮
        ttk.Button(button_frame, text="关闭", command=dialog.destroy).pack(side="right")
        
        # 绑定ESC键关闭对话框
        dialog.bind('<Escape>', lambda e: dialog.destroy())
    
    def create_about_page(self):
        """创建关于页面，显示软件作者信息"""
        # 配置网格权重
        self.about_frame.columnconfigure(0, weight=1)
        self.about_frame.rowconfigure(0, weight=1)
        
        # 主容器
        main_container = ttk.Frame(self.about_frame)
        main_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=20, pady=20)
        main_container.columnconfigure(0, weight=1)
        
        # 软件标题和图标
        title_frame = ttk.Frame(main_container)
        title_frame.grid(row=0, column=0, pady=(0, 30))
        
        # 软件图标（使用emoji代替）
        icon_label = ttk.Label(title_frame, text="📁", font=("Arial", 48))
        icon_label.pack()
        
        # 软件名称
        title_label = ttk.Label(title_frame, text="目录文件扫描器", font=("Arial", 24, "bold"))
        title_label.pack(pady=(10, 5))
        
        # 版本信息
        version_label = ttk.Label(title_frame, text="版本 2.0", font=("Arial", 14))
        version_label.pack()
        
        # 分隔线
        separator = ttk.Separator(main_container, orient="horizontal")
        separator.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=20)
        
        # 软件信息
        info_frame = ttk.Frame(main_container)
        info_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        info_frame.columnconfigure(1, weight=1)
        
        # 软件描述
        desc_text = """一款功能强大的目录文件扫描工具，支持：

• 🔍 智能目录扫描，支持自定义深度和文件类型过滤
• 📊 多种显示模式：树形视图、详细列表、简洁列表
• 📈 详细的文件统计分析和可视化图表
• 📄 多格式导出：Excel（含超链接）、CSV、JSON
• 📝 完整的扫描日志记录
• 🎨 现代化的用户界面设计"""
        
        desc_label = ttk.Label(info_frame, text=desc_text, font=("Arial", 11), justify="left", wraplength=500)
        desc_label.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 20))
        
        # 作者信息
        author_frame = ttk.LabelFrame(info_frame, text="开发信息", padding="15")
        author_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 20))
        author_frame.columnconfigure(1, weight=1)
        
        # 作者
        ttk.Label(author_frame, text="开发者：", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Label(author_frame, text="张牛牛", font=("Arial", 11)).grid(row=0, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        
        # 开发时间
        ttk.Label(author_frame, text="开发时间：", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Label(author_frame, text="2025年7月", font=("Arial", 11)).grid(row=1, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        
        # 技术栈
        ttk.Label(author_frame, text="技术栈：", font=("Arial", 11, "bold")).grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Label(author_frame, text="Python 3.x + Tkinter + Pandas + OpenPyXL", font=("Arial", 11)).grid(row=2, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        
        # 许可证
        ttk.Label(author_frame, text="许可证：", font=("Arial", 11, "bold")).grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Label(author_frame, text="MIT License", font=("Arial", 11)).grid(row=3, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        
        # 联系信息框架
        contact_frame = ttk.LabelFrame(info_frame, text="联系方式", padding="15")
        contact_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 20))
        
        contact_text = "如有问题或建议，欢迎通过以下方式联系：\n\n📧 邮箱：15047831679@126.com"
        ttk.Label(contact_frame, text=contact_text, font=("Arial", 11), justify="left").pack(anchor="w")
        
        # 版权信息
        copyright_label = ttk.Label(main_container, text="© 2025 程序设计与开发：张牛牛. All rights reserved.", 
                                   font=("Arial", 10), foreground="gray")
        copyright_label.grid(row=3, column=0, pady=(20, 0))
        
    def clear_logs(self):
        """清空所有日志"""
        if messagebox.askyesno("确认", "确定要清空所有日志吗？"):
            self.scan_logs.clear()
            self.log_text.configure(state="normal")
            self.log_text.delete(1.0, tk.END)
            self.log_text.configure(state="disabled")
            self.log_info_var.set("日志条数：0")
            self.add_log("日志已清空", "info")
            
    def export_logs(self):
        """导出日志到文件"""
        if not self.scan_logs:
            messagebox.showwarning("警告", "没有日志可导出！")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="保存日志文件",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write("目录文件扫描器 - 扫描日志\n")
                    f.write("=" * 50 + "\n\n")
                    
                    for log in self.scan_logs:
                        f.write(f"[{log['timestamp']}] [{log['type'].upper()}] {log['message']}\n")
                        
                messagebox.showinfo("成功", f"日志已成功导出到：\n{file_path}")
                self.add_log(f"日志已导出到：{file_path}", "success")
            except Exception as e:
                messagebox.showerror("错误", f"导出日志失败：{str(e)}")
                self.add_log(f"导出日志失败：{str(e)}", "error")

def main():
    """主函数"""
    root = tk.Tk()
    app = DirectoryScanner(root)
    root.mainloop()

if __name__ == "__main__":
    main()